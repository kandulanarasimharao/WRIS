import os
import requests
import pandas as pd
from openpyxl import Workbook

BASE = "https://indiawris.gov.in"

HEADERS = {
    "Origin": "https://indiawris.gov.in",
    "Referer": "https://indiawris.gov.in/dataSet/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
    "Content-Type": "application/json"
}

DATASET_CODE = "GWATERLVL"
STATE_NAME = "Andhra Pradesh"
START_DATE = "2000-01-01"
END_DATE = "2025-09-28"

OUTDIR = "WRIS_HTTP_AP"
os.makedirs(OUTDIR, exist_ok=True)


def post(endpoint, payload):
    url = f"{BASE}{endpoint}"
    r = requests.post(url, headers=HEADERS, json=payload)
    r.raise_for_status()
    return r.json()


def save_excel(info, data, fname):
    wb = Workbook()

    # --- Info sheet ---
    ws_info = wb.active
    ws_info.title = "Info"
    for col, (k, v) in enumerate(info.items(), start=1):
        ws_info.cell(row=1, column=col, value=k)
        ws_info.cell(row=2, column=col, value=v)

    # --- Data sheet ---
    ws_data = wb.create_sheet("Data")
    if data:
        keys = list(data[0].keys())
        ws_data.append(keys)
        for row in data:
            ws_data.append([row.get(k, "") for k in keys])

    wb.save(fname)
    print(f"✅ Saved {fname}")


def run():
    # 1. Get states
    states = post("/wris-lapi/StateList", {"datasetcode": DATASET_CODE})["data"]
    state = next(s for s in states if s["stateName"].lower().startswith("andhra"))
    state_code = state["stateCode"]
    print(f"✅ State found: {STATE_NAME} → code {state_code}")

    # 2. Districts
    districts = post("/wris-lapi/getDistrictbyState",
                     {"datasetcode": DATASET_CODE, "statecode": state_code})["data"]

    for d in districts:
        district_name = d["districtname"]

        # 3. Tehsils
        tehsils = post("/wris-lapi/getMasterTehsilList",
                       {"datasetcode": DATASET_CODE, "districtid": d["district_id"]})["data"]

        for t in tehsils:
            tehsil_name = t["tehsilName"]

            # 4. Blocks
            blocks = post("/wris-lapi/getMasterBlockList",
                          {"datasetcode": DATASET_CODE, "tehsilid": t["tehsilid"]})["data"]

            for b in blocks:
                block_name = b["blockName"]

                # 5. Agencies
                agencies = post("/wris-lapi/AgencyListInAnyCase",
                                {"datasetcode": DATASET_CODE,
                                 "blockid": b["blockid"]})["data"]

                for a in agencies:
                    agency_name = a["agencyname"]

                    # 6. Stations
                    stations = post("/stationMaster/getMasterStationsList",
                                    {"stationcode": a["stationcode"],
                                     "datasetcode": DATASET_CODE})["data"]

                    for s in stations:
                        station_code = s["station_Code"]
                        station_name = s["station_Name"]

                        # Metadata
                        info = post("/stationMaster/getMasterStation",
                                    {"stationcode": station_code,
                                     "datasetcode": DATASET_CODE})["data"][0]

                        # Timeseries data
                        data = post("/CommonDataSetMasterAPI/getCommonDataSetByStationCode",
                                    {"station_code": station_code,
                                     "dataset": DATASET_CODE,
                                     "starttime": START_DATE,
                                     "endtime": END_DATE})["data"]

                        # Save Excel
                        safe_station = station_name.replace("/", "-").replace("\\", "-")
                        fname = os.path.join(
                            OUTDIR,
                            f"{STATE_NAME}_{district_name}_{tehsil_name}_{block_name}_{agency_name}_Telemetry_{safe_station}.xlsx"
                        )
                        save_excel(info, data, fname)
                        return  # 🚨 For now, stop after first station (test run)


if __name__ == "__main__":
    run()
